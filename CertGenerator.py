import tkinter as tk # Similar to JPanel, but looks nicer
from tkinter import filedialog # File browsing
import openpyxl # Used to read data from excel files, also the reason I switched to Python
from docx import Document # Used to rewrite data within the Word document
import os

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        excel_file_entry.delete(0, tk.END)
        excel_file_entry.insert(0, file_path)


# Function that uses docx to replaces words in certificates
def generate_certificate(template, name, course_name, field_of_study, date):
    document = Document(template)  #
    # Sets the text that will be replaced with the variables
    for paragraph in document.paragraphs:
        if '{name}' in paragraph.text:
            paragraph.text = paragraph.text.replace('{name}', name)
        if '{course_name}' in paragraph.text:
            paragraph.text = paragraph.text.replace('{course_name}', course_name)
        if '{field_of_study}' in paragraph.text:
            paragraph.text = paragraph.text.replace('{field_of_study}', field_of_study)
        if '{date}' in paragraph.text:
            paragraph.text = paragraph.text.replace('{date}', date)
    document.save(f"certificate_{name}.docx")  # Save the populated document

# Reads Excel workbook and gives a NASBA cert if score is above 3 or non nasba if below 3
def generate_certifications():
    file_path = excel_file_entry.get()
    course_name = course_name_entry.get()
    field_of_study = field_of_study_entry.get()
    date = date_entry.get()
    
    if file_path and course_name and field_of_study and date:
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=2).value
                score = sheet.cell(row=row, column=3).value
                if score is not None and isinstance(score, int):
                    if score >= 3:
                        # IMPORTANT: make sure you change the file path of the templates when using a different computer
                        generate_certificate('F:\\Code\\Python\\NASBATemplate.docx', name, course_name, field_of_study, date,)
                    else:
                        generate_certificate('F:\\Code\\Python\\NonNASBATemplate.docx', name, course_name, field_of_study, date,)
            print("Certificates generated successfully!")
        except Exception as e:
            print(f"Error: {e}")
    else:
        print("Please fill in all the fields.")

def exit_program():
    root.destroy()

def open_help_file():
    try:
        os.system("notepad help.txt")  # Opens the help.txt file in Notepad
    except Exception as e:
        print(f"Error opening help file: {e}")

# Creates the welcome window
welcome_window = tk.Tk()
welcome_window.title("Welcome to Certification Generator")

welcome_label = tk.Label(welcome_window, text="Welcome to the NASBA Certification Generator!")
welcome_label.pack(padx=20, pady=10)

start_button = tk.Button(welcome_window, text="Start", command=welcome_window.destroy)
start_button.pack(padx=20, pady=5)

# Adds a help window that shows documentation upon click
help_button_welcome = tk.Button(welcome_window, text="Help", command=open_help_file)
help_button_welcome.pack(padx=20, pady=5)

welcome_window.mainloop()

# Creates the main window
root = tk.Tk()
root.title("Certification Generator")

# Excel file selection widgets
excel_file_label = tk.Label(root, text="Select Excel File:")
excel_file_label.grid(row=0, column=0, padx=5, pady=5)
excel_file_entry = tk.Entry(root, width=50)
excel_file_entry.grid(row=0, column=1, padx=5, pady=5)
excel_file_button = tk.Button(root, text="Browse", command=select_excel_file)
excel_file_button.grid(row=0, column=2, padx=5, pady=5)

# Course name input
course_name_label = tk.Label(root, text="Course Name:")
course_name_label.grid(row=1, column=0, padx=5, pady=5)
course_name_entry = tk.Entry(root, width=50)
course_name_entry.grid(row=1, column=1, padx=5, pady=5)

# Field of study input
field_of_study_label = tk.Label(root, text="Field of Study:")
field_of_study_label.grid(row=2, column=0, padx=5, pady=5)
field_of_study_entry = tk.Entry(root, width=50)
field_of_study_entry.grid(row=2, column=1, padx=5, pady=5)

# Date input
date_label = tk.Label(root, text="Date:")
date_label.grid(row=3, column=0, padx=5, pady=5)
date_entry = tk.Entry(root, width=50)
date_entry.grid(row=3, column=1, padx=5, pady=5)

# Certification generation button
generate_button = tk.Button(root, text="Generate Certifications", command=generate_certifications)
generate_button.grid(row=4, column=1, padx=5, pady=5)

# Exit button
exit_button = tk.Button(root, text="Exit", command=exit_program)
exit_button.grid(row=5, column=1, padx=5, pady=5)

root.mainloop()
